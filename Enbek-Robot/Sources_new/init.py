import openpyxl
import os
import datetime


def credentials():
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/credentials.xlsx")
    ws = wb.active
    credentials = {}
    for row in range(ws.max_row):
        if str(ws["A" + str(row + 1)].value) == "1CLogin":
            credentials["oneC_login"] = str(ws["B" + str(row + 1)].value)
        if str(ws["A" + str(row + 1)].value) == "1CPassword":
            credentials["oneC_password"] = str(ws["B" + str(row + 1)].value)
        if str(ws["A" + str(row + 1)].value) == "EnbekLogin":
            credentials["enbek_login"] = str(ws["B" + str(row + 1)].value)
        if str(ws["A" + str(row + 1)].value) == "EnbekPassword":
            credentials["enbek_password"] = str(ws["B" + str(row + 1)].value)
        if str(ws["A" + str(row + 1)].value) == "EcpPassword":
            credentials["ecp_password"] = str(ws["B" + str(row + 1)].value)
    return credentials


def period():
    today = datetime.datetime.today().strftime('%d.%m.%Y')

    date = {
        "since": "01.11.2022",
        "till": "30.11.2022"
    }
    return date


def create_logs():
    if not os.path.exists("C:\\Users\\\RobotX\\Desktop\\Ales_enbek\\Robot\\Enbek-Robot\\Tools\\logs\\" + str(datetime.datetime.now().strftime("%d.%m.%Y")) + ".xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Время", "Сотрудник ФИО", "ИИН", "Таб номер", "Статус", "Процесс"])
        wb.save("C:\\Users\\\RobotX\\Desktop\\Ales_enbek\\Robot\\Enbek-Robot\\Tools\\logs\\" + str(
            datetime.datetime.now().strftime("%d.%m.%Y")) + ".xlsx")


def logs(fio, iin, tab_no, status, process):
    create_logs()
    log_path = "C:\\Users\\\RobotX\\Desktop\\Ales_enbek\\Robot\\Enbek-Robot\\Tools\\logs\\" + str(datetime.datetime.now().strftime("%d.%m.%Y")) + ".xlsx"
    wb = openpyxl.load_workbook(log_path)
    ws = wb.active
    ws.append([str(datetime.datetime.now().strftime("%H:%M")), fio, iin, tab_no, status, process])
    wb.save(log_path)


def address_mapping(address):
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/mapping.xlsx")
    ws = wb["Маппинг адресов"]
    address_enbek = {}
    for row in range(ws.max_row):
        if str(ws['B' + str(row + 1)].value).lower().strip() == str(address).lower().strip()[0:3] or str(ws['C' + str(row + 1)].value).lower().strip() == str(address).lower().strip()[0:3]:
            address_enbek = {"obl": str(ws['D' + str(row + 1)].value).split(",")[0].strip(),
                             "city": str(ws['D' + str(row + 1)].value).split(",")[1].strip(),
                             "nas_punkt": str(ws['D' + str(row + 1)].value).split(",")[2].strip()}
            if len(str(ws['D' + str(row + 1)].value).split(",")) > 3:
                address_enbek["street"] = str(ws['D' + str(row + 1)].value).replace(str(ws['D' + str(row + 1)].value).split(",")[0].strip() + str(','), "").replace(str(ws['D' + str(row + 1)].value).split(",")[1].strip() + str(','), "").replace(str(ws['D' + str(row + 1)].value).split(",")[2].strip() + str(','), "").strip()
            break
    if address_enbek == {}:
        address_enbek = {"address": "Не найден в маппинге"}
    return address_enbek


def contract_type_mapping(contract_type):
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/mapping.xlsx")
    ws = wb["Маппинг видов договора"]
    contract_type_enbek = ""
    for row in range(ws.max_row):
        if str(ws['A' + str(row + 1)].value).lower().strip() == str(contract_type).lower().strip():
            contract_type_enbek = str(ws['B' + str(row + 1)].value)
            break
    if contract_type_enbek == "":
        contract_type_enbek = "Не найден в маппинге"
    return contract_type_enbek


def regime_type_mapping(regime_type):
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/mapping.xlsx")
    ws = wb["Маппинг режима работы"]
    regime_type_enbek = ""
    for row in range(ws.max_row):
        if str(ws['A' + str(row + 1)].value).lower().strip() == str(regime_type).lower().strip():
            regime_type_enbek = str(ws['B' + str(row + 1)].value)
            break
    if regime_type_enbek == "":
        regime_type_enbek = "Не найден в маппинге"
    return regime_type_enbek


def position_mapping(position):
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/mapping.xlsx")
    ws = wb["Маппинг должностей"]
    position_enbek = ""
    for row in range(ws.max_row):
        if str(ws['B' + str(row + 1)].value).lower().strip() == str(position).lower().strip():
            position_enbek = str(ws['D' + str(row + 1)].value)
            break
    if position_enbek == "":
        position_enbek = "Не найден в маппинге"
    return position_enbek