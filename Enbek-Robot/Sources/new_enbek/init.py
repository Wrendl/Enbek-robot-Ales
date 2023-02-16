import openpyxl

from Sources.new_enbek.enbek_maps import map_prichina_uvol, map_partTime, location_mapping, map_cato, map_uoz_caption


def fill_contractCate(first_date, second_date):
    if second_date == '':
        return '1'
    elif int(second_date[-2:]) - int(first_date[-2:]) < 2:
        return '2'
    else:
        return '1'


def fill_workingHours(working_hours):
    if working_hours == 'Пятидневка':
        return '1'
    else:
        return '2'


def prof_dict():
    dict = {}
    wb = openpyxl.load_workbook(r"C:\Users\robot.drp\Desktop\Enbek-robot\Tools\Должности.xlsx")
    ws = wb.active
    for row in range(1, ws.max_row):
        dict[str(ws["A" + str(row + 1)].value).lower()] = [ws["B" + str(row + 1)].value, ws["C" + str(row + 1)].value]
    wb.close()
    return dict


def fill_workplace(workplace):
    if workplace['obl'] == 'Алматы':
        workplace['obl'] = 'Г.АЛМАТЫ'
    if workplace['obl'] == 'ГОРОД АЛМАТЫ':
        workplace['obl'] = 'Г.АЛМАТЫ'
    if workplace['obl'] == 'Алматинская область':
        workplace['obl'] = 'АЛМАТИНСКАЯ ОБЛАСТЬ'
    if workplace['city'] == 'Капчагай':
        workplace['city'] = 'Капчагай Г.А.'
        workplace['street'] = 'г.Капчагай'
    return workplace


def prichina_mapping(uvol_arr):
    for uvol in uvol_arr:
        # Причина
        key = str(uvol['Причина расторжения']).split('Кодекса Республики Казахстан')[0] + 'Кодекса Республики Казахстан'
        try:
            uvol['Причина расторжения'] = map_prichina_uvol[key]
        except Exception as e:
            print("нет данной причины увольнения")
            uvol['Причина расторжения'] = map_prichina_uvol['default_value']

        # Дата
        uvol['Дата расторжения'] = uvol['Дата расторжения'][:-2] + '20' + uvol['Дата расторжения'][-2:]
    return uvol_arr


def update_perevod(perevod):
    workplace_values = fill_workplace(perevod['Место выполнения работы'])
    data = {
        'ИИН': perevod['ИИН'],
        'номер допсоглашения': perevod['Номер доп соглашения'],
        'Дата заключения доп.соглашения': perevod['Дата заключения дополнительного соглашения'],
        'Дата начала действия доп.соглашения': perevod['Дата начала действия доп соглашения'],
        'НКЗ': perevod['Должность'],
        'Должность': perevod['Штатная должность'],
        'Регион': workplace_values['obl'],
        'район': workplace_values['city'],
        'Населенный пункт': workplace_values['nas_punkt'],
        'адрес': workplace_values['street'],
        'Срок договора': perevod['Срок договора'],
        'Вид работы': perevod['Вид работы'],
        'Режим рабочего времени': perevod['Режим рабочего времени'],
        'Дата окончания действия дополнительного соглашения': perevod['Дата окончания действия дополнительного соглашения']
    }
    return data
