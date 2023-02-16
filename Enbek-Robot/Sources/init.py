import openpyxl
import datetime
import os
import glob


def logs(arr):
    wb = ""
    ws = ""
    if os.path.exists("C:\\Users\\\RobotX\\Desktop\\Ales_enbek\\Robot\\Enbek-Robot\\Tools\\logs\\" + str(datetime.datetime.now().strftime("%d.%m.%Y")) + ".xlsx"):
        wb = openpyxl.load_workbook("C:\\Users\\\RobotX\\Desktop\\Ales_enbek\\Robot\\Enbek-Robot\\Tools\\logs\\" + str(datetime.datetime.now().strftime("%d.%m.%Y")) + ".xlsx")
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Время", "Сотрудник", "ИИН", "Таб номер", "Статус", "Процесс"])
    ws.append(arr)
    wb.save("C:\\Users\\\RobotX\\Desktop\\Ales_enbek\\Robot\\Enbek-Robot\\Tools\\logs\\" + str(datetime.datetime.now().strftime("%d.%m.%Y")) + ".xlsx")


def get_logs():
    checked_arr = []
    for xls in glob.glob("C:\\Users\\\RobotX\\Desktop\\Ales_enbek\\Robot\\Enbek-Robot\\Tools\\logs\\*"):
        wb = openpyxl.load_workbook(xls)
        ws = wb.active
        for row in range(ws.max_row):
            checked_arr.append({str(ws['B' + str(row+1)].value) + "+" + str(ws['E' + str(row+1)].value): str(ws['C' + str(row+1)].value)})
    return checked_arr


def credentials():
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/credentials.xlsx")
    ws = wb.active
    credentials = {}
    for row in range(ws.max_row):
        if str(ws["A" + str(row + 1)].value) == "1CLogin":
            credentials["oneC_login"] = str(ws["B" + str(row + 1)].value)
        if str(ws["A" + str(row + 1)].value) == "1CPassword":
            credentials["oneC_password"] = str(ws["B" + str(row + 1)].value)
        if str(ws["A" + str(row + 1)].value) == "EnbekLogin":
            credentials["enbek_login"] = str(ws["B" + str(row + 1)].value)
        if str(ws["A" + str(row + 1)].value) == "EnbekPassword":
            credentials["enbek_password"] = str(ws["B" + str(row + 1)].value)
        if str(ws["A" + str(row + 1)].value) == "EcpPassword":
            credentials["ecp_password"] = str(ws["B" + str(row + 1)].value)
    return credentials


def period():
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/Дата.xlsx")
    ws = wb.active
    period = {}
    for row in range(ws.max_row):
        if str(ws["A" + str(row + 1)].value) == "Дата с:":
            period["since"] = str(ws["B" + str(row + 1)].value)
        if str(ws["A" + str(row + 1)].value) == "Дата по:":
            period["till"] = str(ws["B" + str(row + 1)].value)
    if str(period["till"]) == "None":
        period["till"] = " "
    return period


def position_mapping(position):
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/mapping.xlsx")
    ws = wb["Маппинг должностей"]
    position_enbek = ""
    for row in range(ws.max_row):
        if str(ws['B' + str(row + 1)].value).lower().strip() == str(position).lower().strip():
            position_enbek = str(ws['D' + str(row + 1)].value)
            break
    if position_enbek == "":
        position_enbek = "Не найден в маппинге"
    return position_enbek


def address_mapping(address):
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/mapping.xlsx")
    ws = wb["Маппинг адресов"]
    address_enbek = {}
    for row in range(ws.max_row):
        if str(ws['B' + str(row + 1)].value).lower().strip() == str(address).lower().strip()[0:3] or str(ws['C' + str(row + 1)].value).lower().strip() == str(address).lower().strip()[0:3]:
            address_enbek = {"obl": str(ws['D' + str(row + 1)].value).split(",")[0].strip(),
                             "city": str(ws['D' + str(row + 1)].value).split(",")[1].strip(),
                             "nas_punkt": str(ws['D' + str(row + 1)].value).split(",")[2].strip()}
            if len(str(ws['D' + str(row + 1)].value).split(",")) > 3:
                address_enbek["street"] = str(ws['D' + str(row + 1)].value).replace(str(ws['D' + str(row + 1)].value).split(",")[0].strip() + str(','), "").replace(str(ws['D' + str(row + 1)].value).split(",")[1].strip() + str(','), "").replace(str(ws['D' + str(row + 1)].value).split(",")[2].strip() + str(','), "").strip()
            break
    if address_enbek == {}:
        address_enbek = {"address": "Не найден в маппинге"}
    return address_enbek


def contract_type_mapping(contract_type):
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/mapping.xlsx")
    ws = wb["Маппинг видов договора"]
    contract_type_enbek = ""
    for row in range(ws.max_row):
        if str(ws['A' + str(row + 1)].value).lower().strip() == str(contract_type).lower().strip():
            contract_type_enbek = str(ws['B' + str(row + 1)].value)
            break
    if contract_type_enbek == "":
        contract_type_enbek = "Не найден в маппинге"
    return contract_type_enbek


def delete_contract_type_mapping(delete_type):
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/mapping.xlsx")
    ws = wb["Маппинг причин увольнений"]
    delete_type_enbek = ""
    for row in range(ws.max_row):
        if str(ws['A' + str(row + 1)].value).lower().strip() == str(delete_type).lower().strip():
            delete_type_enbek = str(ws['B' + str(row + 1)].value)
            break
    if delete_type_enbek == "":
        delete_type_enbek = "Не найден в маппинге"
    return delete_type_enbek


def regime_type_mapping(regime_type):
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/mapping.xlsx")
    ws = wb["Маппинг режима работы"]
    regime_type_enbek = ""
    for row in range(ws.max_row):
        if str(ws['A' + str(row + 1)].value).lower().strip() == str(regime_type).lower().strip():
            regime_type_enbek = str(ws['B' + str(row + 1)].value)
            break
    if regime_type_enbek == "":
        regime_type_enbek = "Не найден в маппинге"
    return regime_type_enbek


def vacation_mapping(vacation_type):
    wb = openpyxl.load_workbook("C:/Users/RobotX/Desktop/Ales_enbek/Robot/Enbek-Robot/Tools/mapping.xlsx")
    ws = wb["Маппинг отпуска"]
    vacation_type_enbek = ""
    for row in range(ws.max_row):
        if str(ws['A' + str(row + 1)].value).lower().strip() == str(vacation_type).lower().strip():
            vacation_type_enbek = str(ws['B' + str(row + 1)].value)
            break
    if vacation_type_enbek == "":
        vacation_type_enbek = "Не найден в маппинге"
    return vacation_type_enbek
